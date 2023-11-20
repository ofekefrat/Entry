import os
from re import match, sub
from typing import Optional
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from datetime import datetime
from config import *

# TODO implement "class method" construction
# TODO add exceptions for each action, in case it fails
# TODO look into "active cell" in openpyxl to replace passing around row and column

FIRST_POSSIBLE_EMPTY_COLUMN = 2


class Item:
    def __init__(self, serial: str):
        self.wb: Workbook
        self.sheet: Worksheet
        self.row: int
        self.column: int
        self.model_name: str
        self.issuing_date: datetime | str
        self.prev_name: str
        self.is_new: bool
        self.is_returned: bool

        self.serial = serial.strip()
        self.path: str = SERIAL_PATH + self.serial[0:-3] + "000.xlsx"

        try:
            self.wb = load_workbook(self.path)
            self.sheet = self._resolve_sheet()
        except (FileNotFoundError, KeyError) as e:
            raise e

        self.row: int = self._resolve_row()
        self.is_new: bool = self._resolve_is_new()
        if self.is_new:
            self.model_name = ""
            self.column = FIRST_POSSIBLE_EMPTY_COLUMN
        else:
            self._fetch_existing_device_info()
            self.column = self._resolve_column()
            self._fetch_prev_name() # needs some work
            self.is_returned = self._assert_is_returned()

    def _reload(self):
        self.wb = load_workbook(self.path)
        self.sheet = self._resolve_sheet()

    def _resolve_sheet(self) -> Worksheet:
        targetSheet = self.serial[0:-2] + "00"
        return self.wb[targetSheet]

    def _resolve_row(self) -> int:
        # if isinstance(self.sheet, Worksheet):
        cellVal = self.sheet.cell(row=1, column=1).value
        currentRow = 1
        while currentRow < 100 and cellVal != self.serial:
            currentRow += 1
            cellVal = self.sheet.cell(row=currentRow, column=1).value

        if cellVal != self.serial:
            return -1
        else:
            return currentRow

    def _resolve_is_new(self) -> bool:
        return not self._not_last_cell(1)

    def _resolve_column(self) -> int:
        column = self._find_first_empty_cell(FIRST_POSSIBLE_EMPTY_COLUMN)
        result = self._not_last_cell(column)

        while result is not False:
            column = self._find_first_empty_cell(result)
            result = self._not_last_cell(column)

        return column

    def _find_first_empty_cell(self, currentColumn) -> int:
        cellVal = self.sheet.cell(row=self.row, column=currentColumn).value
        while cellVal is not None:
            currentColumn += 1
            cellVal = self.sheet.cell(row=self.row, column=currentColumn).value
        return currentColumn

    def _not_last_cell(self, column):
        lastFound = None
        for i in range(1, 10):
            cellVal = self.sheet.cell(row=self.row, column=(column + i)).value
            if cellVal is not None:
                lastFound = i

        if lastFound is not None:
            return column + lastFound
        else:
            return False

    def _assert_is_returned(self) -> bool:
        cellVal = self.sheet.cell(row=self.row, column=self.column - 1).value
        if type(cellVal) is str:
            return cellVal.strip() == "הוחזר" or cellVal.strip() == "בוטל"
        else:
            return False

    def _fetch_existing_device_info(self):
        model_name = None

        for i in range(4, 7):
            cell = self.sheet.cell(row=self.row, column=i).value
            if model_name is None and type(cell) is str:
                for x in READABLE_MODELS:
                    if x in cell.upper():
                        model_name = cell
                        device_bd = self.sheet.cell(
                            row=self.row, column=i + 1
                        ).value

                        if isinstance(device_bd, datetime):
                            self.issuing_date = device_bd.strftime("%d/%m/%Y")
                        elif type(device_bd) is str:
                            if match(r'^[0-3]?\d/(0|1)?\d/20\d\d$', device_bd): # e.g. 01/02/2003
                                self.issuing_date = datetime.strptime(device_bd, "%d/%m/%y").strftime("%d/%m/%Y")
                            elif match(r'^[0-3]?\d/(0|1)?\d/\d\d$', device_bd): # e.g. 01/02/03
                                self.issuing_date = datetime.strptime(device_bd, "%d/%m/%y").strftime("%d/%m/%Y")
                                # self.issuing_date = deviceBirthday
                            elif match(r'^[0-3]?\d\.(0|1)?\d\.\d\d$', device_bd): # e.g. 01.02.03
                                self.issuing_date = datetime.strptime(device_bd, "%d.%m.%y").strftime("%d/%m/%Y")
                                # self.issuing_date = sub(r'\.', r'/', deviceBirthday)
                            elif match(r'^[0-3]?\d\.(0|1)?\d\.20\d\d$', device_bd): # e.g. 01.02.2003
                                self.issuing_date = datetime.strptime(device_bd, "%d.%m.%Y").strftime("%d/%m/%Y")
                            else:
                                self.issuing_date = "לא תקין"
                        break

        self.model_name = NOT_FOUND if model_name is None else model_name.strip()

    # This one still needs some work, try making it 'return' too
    def _fetch_prev_name(self, sp_column : Optional[int] = None): 
        currentColumn: int = self.column if sp_column is None else sp_column
        cellVal = self.sheet.cell(self.row, currentColumn-1).value
        canceled_flag = True if type(cellVal) is str and cellVal.strip() == "בוטל" else False

        for i in range(4, 8):
            if currentColumn - i >= 1:
                cellVal = self.sheet.cell(self.row, currentColumn - i).value
                if type(cellVal) is str and (
                    cellVal.strip() == "הוחזר" or cellVal.strip() == "בוטל" or cellVal.strip() == self.serial.strip()
                ):
                    if canceled_flag:
                        self._fetch_prev_name(currentColumn-i+1)
                        return
                    try:
                        temp_name = self.sheet.cell(
                            self.row, currentColumn - i + 1
                        ).value
                        
                        if type(temp_name) is str: self.prev_name = temp_name.strip()
                        
                        return
                    except AttributeError as e: # HANDLE!
                        raise e
            else:
                break
        self.prev_name = NOT_FOUND

    def _check_unexpected_entry(self) -> bool:
        self._reload()
        value = self.sheet.cell(self.row, self.column).value
        return value is not None

    # This one might be OK, have a little think about it eh?
    def get_serial_file_permission(self) -> bool:
        self._reload()
        try:
            self.wb.save(self.path)
            return True
        except PermissionError:
            return False
            
    def set_returned(self) -> bool:
        if self._check_unexpected_entry():
            return False
        self.sheet.cell(self.row, self.column).value = "הוחזר"
        self.wb.save(self.path)
        return True

    # this one also needs some thinking
    def update_info(self, name, id, date, model=None):
        if (
            isinstance(self.wb, Workbook)
            and isinstance(self.sheet, Worksheet)
            and isinstance(self.row, int)
            and isinstance(self.column, int)
        ):
            if self._check_unexpected_entry():
                return False
            info = [name, id, model, date]
            for x in info:
                if x != "":
                    self.sheet.cell(self.row, self.column).value = x
                    self.column += 1
            self.wb.save(self.path)
            return True
        return False
