import os.path
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

#TODO everything in getcolumn
#TODO add the option to set items as "returned"
    # TEST: make sure there's no additional cells after the first empty cell found!
    # TEST: serial number validation (insufficient digits, incorrect format...)

class Item:
    def __init__(self, serial):
        self.serial = serial
        self.path = None
        self.wb = None
        self.sheet = None
        self.row = None
        self.column = None
        self.modelName = None
        self.new = None
        self.prevName = None
        self.prevId = None

        self.findWorkbook()
        if not isinstance(self.wb, FileNotFoundError):
            self.findSheet()
            if not isinstance(self.sheet, KeyError):
                self.findRow()
                self.findIsNew()
                if self.new:
                    self.modelName = None
                    self.column = 2
                else:
                    self.findModelName()
                    self.findColumn()
                    self.findPrevInfo()
    

    def createPath(self):
        targetWorkbook = self.serial[0:-3] + "000.xlsx"
        # path = "C:\\Users\\Rachel\\Desktop\\code\\lets\\SerialExcel\\"
        path = "C:\\Users\\ofeke\\vscode projects\\SerialExcel\\"
        self.path = path + targetWorkbook
        

    def findWorkbook(self):
        self.createPath()
        try:
            self.wb = load_workbook(self.path)
        except FileNotFoundError as e:
            self.wb = e
            

    def findSheet(self):
        targetSheet = self.serial[0:-2] + "00"
        try:
            self.sheet = self.wb[targetSheet]   
        except KeyError as e:
            self.sheet = e


    def findRow(self):
        cellVal = self.sheet.cell(row=1, column=1).value
        currentRow=0
        while currentRow < 100 and cellVal != self.serial: 
            currentRow+=1
            cellVal = self.sheet.cell(row=currentRow, column=1).value

        if cellVal != self.serial:
            print("error: serial not found")
            self.row = -1
        
        self.row = currentRow


    def findIsNew (self):
        self.new = not self.not_last_cell(1)


    def findColumn(self):
        # put a flag in the end of the row to signify editing in progress
        # check if there's a flag and throw error if so
        
        column = self.find_next_empty_cell()
        result = self.not_last_cell(column)
        while result is not False:
            column = self.find_next_empty_cell()
            result = self.not_last_cell(column)
            
        cellVal = self.sheet.cell(row=self.row, column=column-1).value
        if cellVal != "הוחזר":
            self.column = -1

        self.column = column


    def findModelName(self):
        models = ["caneo", "domiflex", "exigo", "emineo", "cirrus", "marcus", "f3", "m1", "k300", "pt", "מדרגון", "eloflex", "adiflex"]

        modelName = None
        currentColumn=5
        cellVal1 = self.sheet.cell(row=self.row, column=currentColumn-1).value
        cellVal2 = self.sheet.cell(row=self.row, column=currentColumn).value

        if type(cellVal1) is str:
            for x in models:
                if x in cellVal1.lower():
                    modelName = cellVal1
                    continue
        if modelName is None and type(cellVal2) is str:
            for x in cellVal2.lower():
                if x in cellVal2.lower():
                    modelName = cellVal2
                    continue
        if modelName is None:
            print("model name not found!")

        self.modelName = modelName
        

    def findPrevInfo(self):
        currentColumn = self.column
        for i in range(3, 5): # -1 is "returned", -2 is date
            value = self.sheet.cell(self.row, currentColumn-i).value
            if type(value) is int or value.isnumeric():
                self.prevId = value
                continue
        
        for j in range(0, 2):
            value = self.sheet.cell(self.row, currentColumn-i-j).value
            if type(value) is str and value.isalpha(): # and " " in value:
                self.prevName = value
                continue


    def find_next_empty_cell(self):
        currentColumn=2
        cellVal = self.sheet.cell(row=self.row, column=currentColumn).value
        while cellVal is not None:
            currentColumn+=1
            cellVal = self.sheet.cell(row=self.row, column=currentColumn).value
        return currentColumn


    def not_last_cell(self, column):
        for i in range(1, 10):
            cellVal = self.sheet.cell(row=self.row, column=(column+i)).value
            if cellVal is not None:
                return column+i
            else:
                return False


        

# def find_serial(serial):
#     item = Item(serial)

#     item.getWorkbook()
#     if not isinstance(item.wb, FileNotFoundError):
#         item.getSheet()
#         if not isinstance(item.sheet, KeyError):
#             item.getRow()
#             if item.isNew():
#                 item.new = True
#                 item.modelName = "None"
#                 item.column = 2
#             else:
#                 item.new = False
#                 item.getModelName()
#                 item.getColumn()
#     return item


    def update_info(self, name, id, date, model=None):
        info = [name, id, model, date]
        for x in info:
            if x != "":
                self.sheet.cell(self.row, self.column).value = x
                self.column+=1
        self.wb.save(self.createPath())